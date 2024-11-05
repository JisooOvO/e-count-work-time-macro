from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from datetime import datetime
import os
import time
import pandas as pd
import shutil
import json
import sys
import warnings

warnings.filterwarnings("ignore", category=UserWarning, module='openpyxl.styles.stylesheet')

print("================================")
print("\n")
print("이카운트 근무 시간 자동 계산 매크로")
print("by Jisoo ver 1.0 (2024.11.01)")
print("\n")
print("================================")

if getattr(sys, 'frozen', False):
    script_directory = os.path.dirname(sys.executable)
else:
    script_directory = os.path.dirname(os.path.abspath(__file__))

download_dir = f"{script_directory}/excels" 
json_dir = f"{script_directory}/login_info.json"

if os.path.exists(download_dir):
    shutil.rmtree(download_dir)
    os.makedirs(download_dir)
else :
    os.makedirs(download_dir)

if os.path.exists(json_dir):
    with open(json_dir, 'r') as f :
        login_info = json.load(f)
        com_code = login_info["com_code"]
        user_code = login_info["user_code"]
        password = login_info["password"]
else :
    print("\n")
    print("등록된 계정 정보가 존재하지 않습니다.")
    print("\n")
    print("================================")
    print("\n")
    print("계정 등록을 진행합니다.")
    com_code = input("회사코드 : ")
    user_code = input("아이디 : ")
    password = input("비밀번호 : ")    

    login_info = {
        'com_code': com_code,
        'user_code': user_code,
        'password': password
    }
    
    with open(json_dir, 'w') as f:
        json.dump(login_info, f, ensure_ascii=False, indent=4) 

    print("\n")
    print("계정 정보가 등록되었습니다.")
    print("\n")
    print("================================")

chrome_options = Options()
chrome_options.add_argument("--headless")
chrome_options.add_argument("--no-sandbox")
chrome_options.add_argument("--disable-dev-shm-usage")
chrome_options.add_experimental_option("prefs", {
    "download.default_directory": download_dir, 
    "download.prompt_for_download": False,  
    "download.directory_upgrade": True, 
    "safebrowsing.enabled": True  
})

service = Service(f"{script_directory}/driver/chromedriver")

driver = webdriver.Chrome(service=service, options=chrome_options)

driver.get("https://pay.ecount.com/")

driver.find_element(By.ID, "com_code").send_keys(com_code)
driver.find_element(By.ID, "id").send_keys(user_code)
driver.find_element(By.ID, "passwd").send_keys(password)

driver.find_element(By.ID, "save").click()

WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.ID, "ma9009")))
driver.find_element(By.ID, "ma9009").click()

WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.ID, "searchGroup")))
driver.find_element(By.ID, "searchGroup").click()

WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.ID, "outputExcel")))
driver.find_element(By.ID, "outputExcel").click()

###############################################################################

downloaded_file_path = ''

while len(os.listdir(download_dir)) == 0 : 
    time.sleep(1)

print("\n")
print("엑셀 파일 다운로드 완료")
print("\n")

for file_name in os.listdir(download_dir):
    if file_name.endswith(".xlsx"):
        downloaded_file_path = os.path.join(download_dir, file_name)
        break

data = pd.read_excel(downloaded_file_path, header=1)

if downloaded_file_path:
    date_column = data["일자"]
    working_hours = data["근무시간(시간단위)"]

    base_hours = 9 * 60

    for index, hours in enumerate(working_hours):
        if(date_column[index] == "합계") : break
        if isinstance(hours, str): 
            time_parts = hours.replace("시간", "").replace("분", "").split()
            total_minutes = 0
            total_minutes += int(time_parts[0]) * 60
            
            if(len(time_parts) > 1) :
                total_minutes += int(time_parts[1])
            
            difference = total_minutes - base_hours 

            data.at[index, "근태내역"] = difference            
        else:
            data.at[index, "근태내역"] = 0

    working_times = data["근태내역"]

    for index, times in enumerate(working_times) :
        if(date_column[index] == "합계") : break
        if(index == 0) : data.at[index, "적요"] = times
        else :
            prev_time = data.at[index-1, "적요"]
            if pd.isna(times) :
                data.at[index, "적요"] = 0 + prev_time
            else :
                data.at[index, "적요"] = times + prev_time

data.to_excel(downloaded_file_path, index=False)

###############################################################################

wb = load_workbook(downloaded_file_path)
ws = wb.active

thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))

for index,row in enumerate(ws.iter_rows(max_col=ws.max_column, max_row=ws.max_row)):
    ws.row_dimensions[row[0].row].height = 30
    for cell in row:
        cell.font = Font(size=12)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if(index != 0 and index - 1 < len(date_column)):
            date_str = date_column.iloc[index -1].strip()        
            if len(date_str) == 10 and date_str[4] == '/' and date_str[7] == '/':
                if datetime.strptime(date_str, "%Y/%m/%d").weekday() >= 5 :
                    cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

for column in ws.columns:
    max_length = 0
    column_letter = column[0].column_letter
    for cell in column:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    adjusted_width = (max_length + 10)  
    ws.column_dimensions[column_letter].width = adjusted_width

wb.save(downloaded_file_path)

###############################################################################

os.system(f'open "{downloaded_file_path}"')